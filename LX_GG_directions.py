import pandas as pd
from geocode import get_location
from directions5 import get_optimal_route
import openpyxl
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
import os



# Print iterations progress
def printProgressBar (iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r"):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
        printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
    """
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
    # Print New Line on Complete
    if iteration == total: 
        print()

def getAddressColumn(rng:list, column_endswith:str) -> str:
    for c in rng:
        if c.value.strip().endswith(column_endswith):
            return coordinate_from_string(c.coordinate)[0]
    return 

def do_geocoding(sheetnames: list) -> dict:

    wb = openpyxl.load_workbook(worker_file)  #직원 명부 엑셀파일 로드
    error_messages = []

    if isinstance(sheetnames, str):
        sheetnames = [sheetnames]
    
    for sheetname in sheetnames:
        ws = wb[sheetname]  # 데이터 시트
        max_row = ws.max_row  # 데이터 시트의 마지막 행번호
        max_col = ws.max_column  # 데이터 시트의 마지막 열번호
        

        print(f'[{sheetname}] 시트에 대한 지오코딩을 처리 중입니다.')
        address_col = column_index_from_string(getAddressColumn(ws[1], '주소'))  # 컬럼명이 '주소'로 끝나는 열 번호
        lon_col = column_index_from_string(getAddressColumn(ws[1], 'LON'))  # 컬럼명이 '주소'로 끝나는 열 번호
        lat_col = column_index_from_string(getAddressColumn(ws[1], 'LAT'))  # 컬럼명이 '주소'로 끝나는 열 번호

        for i in range(2, max_row + 1):
            address = ws.cell(i, address_col).value  # 주소값
            lon, lat = ws.cell(i, lon_col).value, ws.cell(i, lat_col).value  # 경, 위도 값

            if not address:  # 주소란에 값이 없으면 패스
                error_messages.append((sheetname, i, "입력 주소값 없음"))
                continue
                
            if lon and lat:  # 기 입력된 경, 위도 값이 있으면 패스
                continue

            lon, lat = get_location(address)

            if not (lon and lat):
                error_messages.append((f'[{sheetname}] 시트', f'{i}행', "주소 변환값을 가져올 수 없음"))

            ws.cell(i, max_col-1).value = lon
            ws.cell(i, max_col).value = lat

            # printProgressBar(i-1, max_row-1)
    
    wb.save(worker_geocode)
    wb.close()
    return {'file': worker_geocode, 'error': error_messages, 'err_count': len(error_messages)}


def calc_distancetime():
    wb = openpyxl.load_workbook(worker_file)  # 직원 명부 엑셀파일 로드
    ws = wb['출퇴근거리']
    worker_ws = wb['직원정보']
    worker_ws.cell(1, ws.max_column+2).value = "작성여부"
    wb.save(worker_geocode)

    worker_df = pd.read_excel(worker_geocode, sheet_name='직원정보', dtype={'사번':str})  # 직원정보를 dataframe으로 가져옴
    jisa_df = pd.read_excel(worker_geocode, sheet_name='지사정보')  # 지사정보를 dataframe으로 가져옴

    # 직원 - 지사간 거리 작성
    for i in range(len(worker_df)):
        checked = worker_df.iloc[i, 6]
        if checked == 'O':
            continue
        try:
            start = worker_df.iloc[i, 4], worker_df.iloc[i, 5]
            worker_info = [v for v in worker_df.iloc[i, 0:2].values]
            for j in range(len(jisa_df)):
                goal = jisa_df.iloc[j, 2], jisa_df.iloc[j, 3]
                jisa_name = jisa_df.iloc[j, 0]
                dist_time = get_optimal_route(start, goal)
                results = worker_info + [jisa_name] + [v for v in dist_time.values()]
                ws.append(results)
                if dist_time['total_duration'] != "Error":
                    worker_ws.cell(i+2, ws.max_column+2).value = "O"

                printProgressBar(i * len(jisa_df) + j + 1, len(worker_df) * len(jisa_df), prefix=worker_info[1])
        except Exception as e:
            print(e)
        
        finally:
            wb.save(worker_geocode)
            wb.close()

if __name__ == '__main__':
    worker_file = '직원명부.xlsx'  #직원 명부 파일 
    worker_geocode = '직원명부_geocode.xlsx'  #직원 명부 파일(지오코딩)

    if os.path.exists(worker_geocode):  
        # 직원명부_geocode.xlsx 파일이 있으면 작업파일을 이것으로 설정
        # 2차, 3차 실행시 좌표변환 결과가 존재하면 건너띄기 위함
        worker_file = worker_geocode

    results = do_geocoding(['직원정보', '지사정보'])
    print('\n',"=" * 22, '지오코딩 결과', "=" * 22 )
    for result in results['error']:
        print(result)

    calc_distancetime()
